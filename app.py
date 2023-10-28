import os
from PIL import Image
import pyocr
import pdf2image #pdfファイルをPIL Imageオブジェクトに変換する
import openpyxl

# pdfファイルをimgオブジェクトに変換
def pdf_to_image(pdf_path):
    img = pdf2image.convert_from_path(pdf_path, 300)
    print(type(img)) #List型
    print(type(img[0])) #PIL.PpmImagePlugin.PpmImageFile型
    return img
    
# 画像から文字を読み込む
def read_text(img):
    builder = pyocr.builders.TextBuilder(tesseract_layout=3)
    text = tool.image_to_string(img[0], lang='jpn', builder=builder)
    return text

# excelファイルを作成
def make_Excel(filename, text):

    # ファイルの存在確認
    if os.path.exists(filename):
        print('このファイルは存在しています')
        book = openpyxl.load_workbook(filename)
    else:
        print('新規ファイルを作成します')
        # ブックを作成
        book = openpyxl.Workbook()

    book_active = book.active
    print(type(book_active)) #openpyxl.worksheet.worksheet.Worksheet型
    book_active.cell(row=1, column=1).value = 'this'
    book.save(filename)
    
if __name__ == '__main__':
    path='C:/Program Files/Tesseract-OCR/tesseract.exe'
    # tesseractパスをセット
    pyocr.tesseract.TESSERACT_CMD = path
    tools = pyocr.get_available_tools()
    tool = tools[0]
    
    # pdfファイルのパスをセット
    pdf_path = 'C:/Users/shodai/Downloads/sample.pdf'
    img = pdf_to_image(pdf_path)
    text = read_text(img)
    make_Excel('sample.xlsx', text)